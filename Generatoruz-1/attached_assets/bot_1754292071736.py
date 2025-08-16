import telebot
from telebot import types
from datetime import datetime , timedelta
import openpyxl
import os
import sqlite3

BOT_TOKEN = "8056968914:AAGmHd9MNUnSSxMJC-mvfztzwB5BuqWeH6Q"
bot = telebot.TeleBot(BOT_TOKEN)

ADMIN_CODE = "1234"
ADMIN_CHAT_ID = 7792775986  # ✅ vergul olib tashlandi

msg_list = []
admin_task_data = {}
employee_states = {}
employee_tasks = {}

employees = {
    "👨‍🔧 Kamol": 7442895800,
    "👨‍🔧 Fozil": 747368650,
    "👨‍🔧 Asomiddin": 1894259641,
    "👨‍🔧 Farruh": 1037206796,
}

def save_task_to_excel(description, location, employees_list, payment, status="⏳ Davom etmoqda"):
    file_name = "topshiriqlar.xlsx"
    headers = ["Sana", "Vaqt", "Topshiriq", "Lokatsiya", "Hodimlar", "Pul miqdori", "Holat"]

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(headers)
    else:
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

    loc_str = f"{location.latitude}, {location.longitude}"
    emp_str = ", ".join(employees_list)

    sheet.append([date_str, time_str, description, loc_str, emp_str, payment, status])
    wb.save(file_name)

@bot.message_handler(commands=['start'])
def start_message(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("🔐 Admin", "👤 Xodim")
    bot.send_message(message.chat.id, "Assalomu alaykum!\nIltimos, rolingizni tanlang:", reply_markup=markup)

@bot.message_handler(commands=['getid'])
def send_admin_id(message):
    bot.reply_to(message, f"🆔 Sizning chat ID'ingiz: {message.chat.id}")

@bot.message_handler(func=lambda message: message.text == "🔐 Admin")
def ask_code(message):
    msg = bot.send_message(message.chat.id, "🔑 Iltimos, admin kodini kiriting:")
    bot.register_next_step_handler(msg, verify_code)

def verify_code(message):
    if message.text == ADMIN_CODE:
        bot.send_message(message.chat.id, "✅ Xush kelibsiz, admin!")
        show_admin_panel(message)
    else:
        msg = bot.send_message(message.chat.id, "❌ Kod noto‘g‘ri. Qaytadan urinib ko‘ring:")
        bot.register_next_step_handler(msg, verify_code)

def show_admin_panel(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("📝 Topshiriqlar", "📊 Qarzdorlik", "📋 Hisobot", "⬅️ Ortga")
    msg = bot.send_message(message.chat.id, "Admin panelga xush kelibsiz!", reply_markup=markup)
    msg_list.append(msg)

def delete_old_messages():
    for msg in msg_list:
        try:
            bot.delete_message(msg.chat.id, msg.message_id)
        except Exception as e:
            print(f"Xabar o‘chirilmadi: {e}")
    msg_list.clear()


    # Har bir topshiriqni alohida yuborish:
    for msg in msg_list:
        bot.send_message(message.chat.id, msg)


@bot.message_handler(func=lambda message: message.text == "📝 Topshiriqlar")
def start_task_creation(message):
    chat_id = message.chat.id
    admin_task_data[chat_id] = {}
    msg = bot.send_message(chat_id, "📝 Topshiriq matnini kiriting:")
    bot.register_next_step_handler(msg, get_task_text)

def get_task_text(message):
    chat_id = message.chat.id
    admin_task_data[chat_id]["description"] = message.text

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    loc_btn = types.KeyboardButton("📍 Lokatsiyani yuborish", request_location=True)
    markup.add(loc_btn)
    bot.send_message(chat_id, "📍 Lokatsiyani yuboring:", reply_markup=markup)

@bot.message_handler(content_types=['location'])
def receive_location(message):
    chat_id = message.chat.id
    lat = message.location.latitude
    lon = message.location.longitude
    admin_task_data[chat_id]["location"] = message.location

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("💰 Pul miqdori")
    bot.send_message(chat_id, "✅ Lokatsiya qabul qilindi.\n💰 Endi pul miqdorini kiriting:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "💰 Pul miqdori")
def ask_payment(message):
    msg = bot.send_message(message.chat.id, "💸 Pul miqdorini kiriting:")
    bot.register_next_step_handler(msg, save_payment)

def save_payment(message):
    chat_id = message.chat.id
    admin_task_data[chat_id]["payment"] = message.text

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("👥 Kerakli hodimlar")
    bot.send_message(chat_id, "👥 Endi kerakli hodimlarni tanlang:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "👥 Kerakli hodimlar")
def choose_employees(message):
    chat_id = message.chat.id
    admin_task_data[chat_id]["selected"] = []

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    for name in employees:
        markup.add(name)
    markup.add("📨 Yuborish")
    bot.send_message(chat_id, "Tanlang (bir nechta hodim tanlashingiz mumkin):", reply_markup=markup)
    bot.register_next_step_handler(message, select_employee)

def select_employee(message):
    chat_id = message.chat.id
    name = message.text

    if name == "📨 Yuborish":
        send_task_to_employees(message)
        return

    if name in employees:
        if name not in admin_task_data.get(chat_id, {}).get("selected", []):
            admin_task_data[chat_id]["selected"].append(name)
            bot.send_message(chat_id, f"✅ {name} tanlandi.")
        else:
            bot.send_message(chat_id, f"⚠️ {name} allaqachon tanlangan.")
    else:
        bot.send_message(chat_id, "❌ Tugmalardan birini tanlang.")

    bot.register_next_step_handler(message, select_employee)

def send_task_to_employees(message):
    chat_id = message.chat.id
    data = admin_task_data.get(chat_id)

    if not data or "location" not in data or "description" not in data or "payment" not in data or not data.get("selected"):
        bot.send_message(chat_id, "❌ Ma'lumotlar to‘liq emas. Iltimos, qaytadan boshlang.")
        return

    lat, lon = data["location"].latitude, data["location"].longitude
    task_text = f"📢 Sizga yangi topshiriq:\n\n📝 {data['description']}\n📍 Lokatsiya: xaritada\n💰 Pul: {data['payment']} so‘m"

    for name in data["selected"]:
        user_id = employees.get(name)
        if not user_id:
            continue
        try:
            bot.send_message(user_id, task_text)
            bot.send_location(user_id, latitude=lat, longitude=lon)

            # ✅ Hodimga topshiriq qo‘shish
            short_name = name.split()[-1]
            employee_tasks.setdefault(short_name, []).append({
                'description': data['description'],
                'location': f"{lat}, {lon}",
                'payment': data['payment'],
                'phone': "Telefon yo‘q",
                'status': "⏳ Davom etmoqda"
            
                
            })

        except Exception as e:
            bot.send_message(chat_id, f"⚠️ {name} ga yuborilmadi.\nXato: {e}")

    save_task_to_excel(data['description'], data['location'], data['selected'], data['payment'])
    bot.send_message(chat_id, "✅ Topshiriq yuborildi. Bajarilishini kuting.")
    show_admin_panel(message)

@bot.message_handler(func=lambda message: message.text == "📋 Hisobot")
def show_report_menu(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("📅 30 kunlik hisobot", "🗓 1 haftalik hisobot")
    markup.add("📤 Excel faylga chop etish", "🔙 Ortga")
    bot.send_message(message.chat.id, "Quyidagilardan birini tanlang:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "📅 30 kunlik hisobot")
def report_30_days(message):
    chat_id = message.chat.id
    name = employee_states.get(chat_id, {}).get("name", "Noma'lum")
    today = datetime.now()
    start_date = today - timedelta(days=30)

    conn = sqlite3.connect("tasks.db")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT vazifa, manzil, summa, sana FROM tasks
        WHERE xodim = ? AND status = 'Bajarildi' AND sana BETWEEN ? AND ?
    """, (name, start_date.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")))
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        bot.send_message(chat_id, "Oxirgi 30 kunda bajarilgan vazifalar yo‘q.")
    else:
        total = sum(row[2] for row in rows)
        bot.send_message(chat_id, f"✅ 30 kun ichida {len(rows)} ta vazifa bajarilgan.\n💰 Umumiy to‘lov: {total} so‘m.")

@bot.message_handler(func=lambda message: message.text == "🗓 1 haftalik hisobot")
def report_7_days(message):
    chat_id = message.chat.id
    name = employee_states.get(chat_id, {}).get("name", "Noma'lum")
    today = datetime.now()
    start_date = today - timedelta(days=7)

    conn = sqlite3.connect("tasks.db")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT vazifa, manzil, summa, sana FROM tasks
        WHERE xodim = ? AND status = 'Bajarildi' AND sana BETWEEN ? AND ?
    """, (name, start_date.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")))
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        bot.send_message(chat_id, "Oxirgi 7 kunda bajarilgan vazifalar yo‘q.")
    else:
        total = sum(row[2] for row in rows)
        bot.send_message(chat_id, f"✅ 1 hafta ichida {len(rows)} ta vazifa bajarilgan.\n💰 Umumiy to‘lov: {total} so‘m.")


@bot.message_handler(func=lambda message: message.text == "📤 Excel faylga chop etish")
def export_excel_report(message):
    chat_id = message.chat.id
    name = employee_states.get(chat_id, {}).get("name", "Noma'lum")
    today = datetime.now()
    start_date = today - timedelta(days=30)

    conn = sqlite3.connect("tasks.db")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT vazifa, manzil, summa, sana FROM tasks
        WHERE xodim = ? AND status = 'Bajarildi' AND sana BETWEEN ? AND ?
    """, (name, start_date.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")))
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        bot.send_message(chat_id, "Excelga chop etiladigan vazifalar yo‘q.")
        return

    filename = f"{name}_hisobot_{today.strftime('%Y%m%d')}.xlsx"
    path = os.path.join("hisobotlar", filename)
    os.makedirs("hisobotlar", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Vazifa", "Manzil", "Summa", "Sana"])

    for row in rows:
        ws.append(row)

    wb.save(path)
    with open(path, "rb") as f:
        bot.send_document(chat_id, f, caption="📤 Excel hisobotingiz tayyor!")

@bot.message_handler(func=lambda message: message.text == "🔙 Ortga")
def go_back(message):
    # Xodim paneliga qaytish kodi shu yerda bo‘lishi kerak
    show_employee_panel(message)

import sqlite3
from telebot import types

# Baza yaratish
def init_debt_db():
    conn = sqlite3.connect("qarzdorlik.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS debts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_name TEXT,
            amount REAL,
            reason TEXT,
            date TEXT
        )
    """)
    conn.commit()
    conn.close()

# Qarz qo‘shish
def add_debt(employee_name, amount, reason, date):
    conn = sqlite3.connect("qarzdorlik.db")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO debts (employee_name, amount, reason, date) VALUES (?, ?, ?, ?)",
                   (employee_name, amount, reason, date))
    conn.commit()
    conn.close()

# Qarzlarni ko‘rish
def get_debts_by_employee(employee_name):
    conn = sqlite3.connect("qarzdorlik.db")
    cursor = conn.cursor()
    cursor.execute("SELECT amount, reason, date FROM debts WHERE employee_name=?", (employee_name,))
    rows = cursor.fetchall()
    conn.close()
    return rows

# Admin tugmasi orqali qarzlarni chiqarish
@bot.message_handler(func=lambda message: message.text == "📊 Qarzdorlik")
def show_debt_menu(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("➕ Qarz qo‘shish", "📋 Qarzlarni ko‘rish", "🔙 Orqaga")
    bot.send_message(message.chat.id, "Qarzdorlik bo‘limi:", reply_markup=markup)

# Qarz qo‘shish bosqichi
@bot.message_handler(func=lambda message: message.text == "➕ Qarz qo‘shish")
def start_adding_debt(message):
    bot.send_message(message.chat.id, "Xodim ismini kiriting:")
    bot.register_next_step_handler(message, get_debt_name)

def get_debt_name(message):
    employee_name = message.text
    bot.send_message(message.chat.id, f"{employee_name} uchun qarz summasini kiriting:")
    bot.register_next_step_handler(message, lambda msg: get_debt_amount(msg, employee_name))

def get_debt_amount(message, employee_name):
    try:
        amount = float(message.text)
        bot.send_message(message.chat.id, "Qarz sababini yozing:")
        bot.register_next_step_handler(message, lambda msg: get_debt_reason(msg, employee_name, amount))
    except ValueError:
        bot.send_message(message.chat.id, "Iltimos, faqat raqam kiriting!")

def get_debt_reason(message, employee_name, amount):
    reason = message.text
    from datetime import datetime
    date = datetime.now().strftime("%Y-%m-%d")
    add_debt(employee_name, amount, reason, date)
    bot.send_message(message.chat.id, f"✅ {employee_name} uchun {amount} so‘m qarz qo‘shildi.")

# Qarzlarni ko‘rish
@bot.message_handler(func=lambda message: message.text == "📋 Qarzlarni ko‘rish")
def view_debts(message):
    bot.send_message(message.chat.id, "Xodim ismini kiriting:")
    bot.register_next_step_handler(message, show_debt_for_employee)

def show_debt_for_employee(message):
    employee_name = message.text
    debts = get_debts_by_employee(employee_name)
    if not debts:
        bot.send_message(message.chat.id, f"{employee_name} uchun qarz topilmadi.")
        return
    msg = f"📋 {employee_name} uchun qarzdorlik ro‘yxati:\n\n"
    for amount, reason, date in debts:
        msg += f"💰 {amount} so‘m | 📝 {reason} | 📅 {date}\n"
    bot.send_message(message.chat.id, msg)

# Orqaga qaytish
@bot.message_handler(func=lambda message: message.text == "🔙 Orqaga")
def go_back(message):
    from_panel = types.ReplyKeyboardMarkup(resize_keyboard=True)
    from_panel.add("🏠 Asosiy menyu")
    bot.send_message(message.chat.id, "Asosiy menyuga qaytdingiz.", reply_markup=from_panel)


# 👤 Xodim login
@bot.message_handler(func=lambda message: message.text == "👤 Xodim")
def start_employee(message):
    bot.send_message(message.chat.id, "👤 Ismingizni kiriting:")
    employee_states[message.chat.id] = {'step': 'ask_name'}

@bot.message_handler(func=lambda message: employee_states.get(message.chat.id, {}).get('step') == 'ask_name')
def step_name(message):
    employee_states[message.chat.id] = {'step': 'ask_password', 'name': message.text}
    bot.send_message(message.chat.id, "🔐 Maxfiy parolingizni kiriting:")

@bot.message_handler(func=lambda message: employee_states.get(message.chat.id, {}).get('step') == 'ask_password')
def step_password(message):
    chat_id = message.chat.id
    name = employee_states[chat_id]['name']
    entered_password = message.text

    employee_passwords = {
        "Kamol": "1234",
        "Fozil": "5678",
        "Asomiddin": "abcd",
        "Farruh": "f455",
    }

    if employee_passwords.get(name) == entered_password:
        employee_states[chat_id] = {'name': name, 'step': None}
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(types.InlineKeyboardButton("📝 Topshiriqlar", callback_data="view_tasks"))
        keyboard.add(types.InlineKeyboardButton("💰 Hisob", callback_data="view_balance"))
        keyboard.add(types.InlineKeyboardButton("📜 Tarix", callback_data="task_history"))
        bot.send_message(chat_id, f"Xush kelibsiz, {name}!", reply_markup=keyboard)
    else:
        bot.send_message(chat_id, "❌ Notog‘ri parol! Qaytadan urinib ko‘ring.")

# 👨‍🔧 Xodim paneli tugmalari

@bot.callback_query_handler(func=lambda call: call.data.startswith("accept_"))
def accept_task(call):
    name = call.data.split("_")[1]
    chat_id = call.message.chat.id
    bot.answer_callback_query(call.id)

    if name in employee_tasks:
        for task in reversed(employee_tasks[name]):
            if task['status'] == "Yuborilgan":
                task['status'] = "Qabul qilingan"
                break


@bot.callback_query_handler(func=lambda call: call.data.startswith("complete_"))
def complete_task(call):
    name = call.data.split("_")[1]
    chat_id = call.message.chat.id
    employee_states[chat_id] = {
        "step": "ask_desc",
        "name": name
    }
    bot.send_message(chat_id, "📝 Topshiriq bajarganingiz haqida tavsif yuboring:")

@bot.message_handler(func=lambda m: employee_states.get(m.chat.id, {}).get("step") == "ask_desc")
def get_desc(m):
    employee_states[m.chat.id]["desc"] = m.text
    employee_states[m.chat.id]["step"] = "ask_photo"
    bot.send_message(m.chat.id, "📷 Endi topshiriqdan rasm yuboring:")

@bot.message_handler(content_types=['photo'])
def get_photo(m):
    if employee_states.get(m.chat.id, {}).get("step") != "ask_photo":
        return
    name = employee_states[m.chat.id]["name"]
    desc = employee_states[m.chat.id]["desc"]
    photo = m.photo[-1].file_id

    # Topshiriqni bajarilgan deb belgilash
    for task in reversed(employee_tasks.get(name, [])):
        if task["status"] in ["Qabul qilingan", "Yuborilgan"]:
            task["status"] = "✅ Bajarildi"
            break

    text = f"📥 *Yangi xabar*\n\n" \
           f"👷‍♂️ {name} topshiriqni bajardi.\n" \
           f"📝 {desc}\n" \
           f"💰 {task['payment']} so‘m"
    bot.send_photo(ADMIN_CHAT_ID, photo, caption=text, parse_mode='Markdown')
    bot.send_message(m.chat.id, "✅ Rahmat! Ma'lumot yuborildi.")
    employee_states[m.chat.id] = {}

# ADMIN: 24 soat ichidagi xabarlar
@bot.message_handler(func=lambda msg: msg.text == "📩 Xabarlar")
def show_recent_reports(msg):
    now = datetime.now()
    one_day_ago = now -timedelta(days=1)
    found = False
    for name, tasks in employee_tasks.items():
        for task in tasks:
            if task.get("status") == "✅ Bajarildi" and task["time"] >= one_day_ago:
                text = f"👷‍♂️ {name} | 📝 {task['description']} | 💰 {task['payment']} so‘m"
                bot.send_message(msg.chat.id, text)
                found = True
    if not found:
        bot.send_message(msg.chat.id, "⏱️ So‘nggi 24 soatda hech qanday topshiriq bajarilmagan.")

@bot.callback_query_handler(func=lambda call: call.data == "view_tasks")
def view_tasks(call):
    chat_id = call.message.chat.id
    name = employee_states.get(chat_id, {}).get("name")
    if not name:
        return bot.send_message(chat_id, "Avval tizimga kiring.")
    tasks = employee_tasks.get(name, [])
    if not tasks:
        return bot.send_message(chat_id, "Sizda hozircha topshiriq yo‘q.")
    for i, task in enumerate(tasks):
        msg = (
            f"📌 *Topshiriq:*\n{task['description']}\n\n"
            f"📍 *Joy:* {task['location']}\n"
            f"📞 *Tel:* {task['phone']}\n"
            f"💰 *Pul:* {task['payment']} so‘m"
        )
        keyboard = types.InlineKeyboardMarkup()
        keyboard.add(
            types.InlineKeyboardButton("✅ Bajarildi", callback_data=f"task_done_{i}"),
            types.InlineKeyboardButton("❌ Bajarilmadi", callback_data=f"task_fail_{i}")
        )
        bot.send_message(chat_id, msg, parse_mode='Markdown', reply_markup=keyboard)

@bot.callback_query_handler(func=lambda call: call.data.startswith("task_done_"))
def done_task(call):
    index = int(call.data.split("_")[-1])
    chat_id = call.message.chat.id
    employee_states[chat_id]['step'] = 'task_done_desc'
    employee_states[chat_id]['task_index'] = index
    bot.send_message(chat_id, "📝 Topshiriq bajarganingiz haqida qisqacha tavsif yozing:")

@bot.message_handler(func=lambda message: employee_states.get(message.chat.id, {}).get('step') == 'task_done_desc')
def get_task_desc(message):
    chat_id = message.chat.id
    employee_states[chat_id]['task_desc'] = message.text
    employee_states[chat_id]['step'] = 'task_done_photo'
    bot.send_message(chat_id, "📷 Topshiriqdan rasm yuboring:")

@bot.message_handler(content_types=['photo'])
def get_task_photo(message):
    chat_id = message.chat.id
    if employee_states.get(chat_id, {}).get('step') != 'task_done_photo':
        return
    photo = message.photo[-1].file_id
    desc = employee_states[chat_id]['task_desc']
    index = employee_states[chat_id]['task_index']
    name = employee_states[chat_id]['name']
    task = employee_tasks[name][index]
    task['status'] = "✅ Bajarildi"
    msg = (
        f"👷‍♂️ *{name}* topshiriqni bajardi!\n"
        f"📝 Tavsif: {desc}\n"
        f"📍 Joy: {task['location']}\n"
        f"💰 Pul: {task['payment']} so‘m\n"
        f"📞 Tel: {task['phone']}"
    )
    bot.send_photo(ADMIN_CHAT_ID, photo, caption=msg, parse_mode='Markdown')
    bot.send_message(chat_id, "✅ Rahmat! Ma'lumotlar yuborildi.")
    employee_states[chat_id]['step'] = None

@bot.callback_query_handler(func=lambda call: call.data.startswith("task_fail_"))
def fail_task(call):
    index = int(call.data.split("_")[-1])
    chat_id = call.message.chat.id
    name = employee_states.get(chat_id, {}).get("name")
    if not name:
        return
    employee_tasks[name][index]['status'] = "❌ Bajarilmadi"
    bot.send_message(chat_id, "❌ Topshiriq bajarilmagan deb belgilandi.")

@bot.callback_query_handler(func=lambda call: call.data == "view_balance")
def show_balance(call):
    name = employee_states.get(call.message.chat.id, {}).get("name")
    if not name:
        return
    total = sum(int(task['payment']) for task in employee_tasks.get(name, []) if task.get('status') == "✅ Bajarildi")
    bot.send_message(call.message.chat.id, f"💵 Jami daromadingiz: {total} so‘m")

@bot.callback_query_handler(func=lambda call: call.data == "task_history")
def history(call):
    name = employee_states.get(call.message.chat.id, {}).get("name")
    if not name:
        return
    messages = []
    for task in employee_tasks.get(name, []):
        status = task.get('status', '⏳ Davom etmoqda')
        messages.append(
            f"📝 {task['description']} | 💰 {task['payment']} | 📍 {task['location']} | {status}"
        )
    text = "\n\n".join(messages) if messages else "Sizda hali tarix yo‘q."
    bot.send_message(call.message.chat.id, text)

print("🤖 Bot ishga tushdi...")
bot.infinity_polling()
