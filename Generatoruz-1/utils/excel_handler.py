import openpyxl
import os
from datetime import datetime
from typing import List, Dict

class ExcelHandler:
    def __init__(self, file_path: str = "topshiriqlar.xlsx"):
        self.file_path = file_path
    
    def save_task_to_excel(self, description: str, location, employees_list: List[str], 
                          payment: str, status: str = "⏳ Davom etmoqda"):
        """Save task to Excel file"""
        headers = ["Sana", "Vaqt", "Topshiriq", "Lokatsiya", "Hodimlar", "Pul miqdori", "Holat"]
        
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        # Create or load workbook
        if not os.path.exists(self.file_path):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(headers)
        else:
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active
        
        # Format location and employees
        if hasattr(location, 'latitude') and hasattr(location, 'longitude'):
            loc_str = f"{location.latitude}, {location.longitude}"
        else:
            loc_str = str(location)
        
        emp_str = ", ".join(employees_list)
        
        # Add row
        sheet.append([date_str, time_str, description, loc_str, emp_str, payment, status])
        wb.save(self.file_path)
    
    def create_employee_report(self, employee_name: str, tasks: List[Dict]) -> str:
        """Create Excel report for employee"""
        if not tasks:
            return None
        
        # Create reports directory
        os.makedirs("hisobotlar", exist_ok=True)
        
        # Generate filename
        today = datetime.now()
        filename = f"{employee_name}_hisobot_{today.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join("hisobotlar", filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{employee_name} Hisoboti"
        
        # Add headers
        headers = ["Vazifa", "Manzil", "Summa", "Sana"]
        ws.append(headers)
        
        # Add data
        total_amount = 0
        for task in tasks:
            ws.append([
                task.get('description', ''),
                task.get('location', ''),
                task.get('amount', 0),
                task.get('date', '')
            ])
            total_amount += task.get('amount', 0)
        
        # Add total row
        ws.append([])
        ws.append(["JAMI:", "", total_amount, ""])
        
        # Save file
        wb.save(filepath)
        return filepath
    
    def create_debt_report(self, employee_name: str, debts: List[Dict]) -> str:
        """Create debt report for employee"""
        if not debts:
            return None
        
        # Create reports directory
        os.makedirs("hisobotlar", exist_ok=True)
        
        # Generate filename
        today = datetime.now()
        filename = f"{employee_name}_qarz_hisoboti_{today.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join("hisobotlar", filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{employee_name} Qarz Hisoboti"
        
        # Add headers
        headers = ["Summa", "Sabab", "Sana"]
        ws.append(headers)
        
        # Add data
        total_debt = 0
        for debt in debts:
            ws.append([
                debt.get('amount', 0),
                debt.get('reason', ''),
                debt.get('date', '')
            ])
            total_debt += debt.get('amount', 0)
        
        # Add total row
        ws.append([])
        ws.append(["JAMI QARZ:", "", total_debt])
        
        # Save file
        wb.save(filepath)
        return filepath
