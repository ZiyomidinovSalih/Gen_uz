import os
import json
import openpyxl
from datetime import datetime, timedelta
from typing import List, Tuple, Optional, Dict, Any
from config import REPORTS_DIR, MEDIA_DIR
from database import get_employee_tasks, get_debts, get_task_statistics

def ensure_directories():
    """Ensure required directories exist"""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    os.makedirs(MEDIA_DIR, exist_ok=True)

def save_media_file(file_info, bot, media_type: str) -> str:
    """Save media file and return file path"""
    ensure_directories()
    
    # Download file
    downloaded_file = bot.download_file(file_info.file_path)
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_extension = os.path.splitext(file_info.file_path)[1]
    filename = f"{media_type}_{timestamp}{file_extension}"
    
    # Save file
    filepath = os.path.join(MEDIA_DIR, filename)
    with open(filepath, 'wb') as f:
        f.write(downloaded_file)
    
    return filepath

def format_task_info(task: Tuple) -> str:
    """Format task information for display"""
    (task_id, description, location_lat, location_lon, location_address,
     payment_amount, assigned_to, assigned_by, status, created_at,
     started_at, completed_at, completion_report, completion_media, received_amount) = task
    
    # Format creation time
    try:
        created_time = datetime.fromisoformat(created_at).strftime("%d.%m.%Y %H:%M")
    except:
        created_time = created_at
    
    # Status emoji
    status_emoji = {
        "pending": "⏳",
        "in_progress": "🔄",
        "completed": "✅"
    }.get(status, "❓")
    
    # Safe status formatting
    status_text = status.title() if status else "Noma'lum"
    
    task_text = f"""
🆔 Vazifa ID: {task_id}
{status_emoji} Holat: {status_text}

📝 Tavsif: {description}
💰 To'lov: {f"{payment_amount:,.0f} so'm" if payment_amount else "Belgilanmagan"}
📅 Yaratilgan: {created_time}
"""
    
    if location_lat and location_lon:
        task_text += f"📍 Lokatsiya: {location_lat:.6f}, {location_lon:.6f}\n"
    
    if status == "completed" and received_amount is not None:
        task_text += f"💵 Olingan: {received_amount:,.0f} so'm\n"
    
    if completion_report:
        task_text += f"📋 Hisobot: {completion_report[:100]}{'...' if len(completion_report) > 100 else ''}\n"
    
    return task_text.strip()

def generate_employee_report(employee_name: str, days: int = 30) -> Optional[str]:
    """Generate Excel report for employee"""
    ensure_directories()
    
    # Get completed tasks
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    tasks = get_employee_tasks(employee_name, "completed")
    if not tasks:
        return None
    
    # Filter by date
    filtered_tasks = []
    for task in tasks:
        try:
            task_date = datetime.fromisoformat(task[9])  # created_at
            if start_date <= task_date <= end_date:
                filtered_tasks.append(task)
        except:
            continue
    
    if not filtered_tasks:
        return None
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{employee_name} Hisoboti"
    
    # Headers
    headers = [
        "ID", "Tavsif", "To'lov (so'm)", "Olingan (so'm)", 
        "Yaratilgan", "Yakunlangan", "Hisobot"
    ]
    ws.append(headers)
    
    # Data
    total_payment = 0
    total_received = 0
    
    for task in filtered_tasks:
        (task_id, description, location_lat, location_lon, location_address,
         payment_amount, assigned_to, assigned_by, status, created_at,
         started_at, completed_at, completion_report, completion_media, received_amount) = task
        
        total_payment += payment_amount or 0
        total_received += received_amount or 0
        
        # Format dates
        try:
            created_formatted = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
        except:
            created_formatted = created_at
        
        try:
            completed_formatted = datetime.fromisoformat(completed_at).strftime("%d.%m.%Y") if completed_at else ""
        except:
            completed_formatted = completed_at or ""
        
        ws.append([
            task_id,
            description[:50] + ("..." if len(description) > 50 else ""),
            payment_amount or 0,
            received_amount or 0,
            created_formatted,
            completed_formatted,
            (completion_report[:30] + "...") if completion_report and len(completion_report) > 30 else (completion_report or "")
        ])
    
    # Summary
    ws.append([])
    ws.append(["JAMI:", "", total_payment, total_received, "", "", ""])
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{employee_name}_hisobot_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    wb.save(filepath)
    return filepath

def generate_admin_report() -> Optional[str]:
    """Generate comprehensive admin report"""
    ensure_directories()
    
    # Get statistics
    stats = get_task_statistics()
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Tasks summary sheet
    ws1 = wb.active
    ws1.title = "Umumiy ma'lumot"
    
    ws1.append(["Vazifalar statistikasi"])
    ws1.append([])
    ws1.append(["Jami vazifalar:", stats["total_tasks"]])
    ws1.append(["Jami to'lovlar:", f"{stats['total_payments']:,.0f} so'm"])
    ws1.append(["Jami qarzlar:", f"{stats['total_debts']:,.0f} so'm"])
    ws1.append([])
    
    ws1.append(["Holat bo'yicha:"])
    for status, count in stats["status_counts"].items():
        status_name = {
            "pending": "Kutilayotgan",
            "in_progress": "Bajarilmoqda",
            "completed": "Yakunlangan"
        }.get(status, status)
        ws1.append([status_name, count])
    
    # Tasks details sheet
    ws2 = wb.create_sheet("Vazifalar")
    from database import DATABASE_PATH
    import sqlite3
    
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, description, assigned_to, payment_amount, received_amount, 
               status, created_at, completed_at
        FROM tasks
        ORDER BY created_at DESC
    """)
    
    tasks_data = cursor.fetchall()
    conn.close()
    
    # Headers
    ws2.append(["ID", "Tavsif", "Xodim", "To'lov", "Olingan", "Holat", "Yaratilgan", "Yakunlangan"])
    
    for task in tasks_data:
        task_id, description, assigned_to, payment_amount, received_amount, status, created_at, completed_at = task
        
        # Format dates
        try:
            created_formatted = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
        except:
            created_formatted = created_at
        
        try:
            completed_formatted = datetime.fromisoformat(completed_at).strftime("%d.%m.%Y") if completed_at else ""
        except:
            completed_formatted = ""
        
        status_name = {
            "pending": "Kutilayotgan",
            "in_progress": "Bajarilmoqda", 
            "completed": "Yakunlangan"
        }.get(status, status)
        
        ws2.append([
            task_id,
            description[:50] + ("..." if len(description) > 50 else ""),
            assigned_to,
            payment_amount,
            received_amount or 0,
            status_name,
            created_formatted,
            completed_formatted
        ])
    
    # Debts sheet
    ws3 = wb.create_sheet("Qarzlar")
    debts = get_debts()
    
    ws3.append(["ID", "Xodim", "Miqdor", "Sabab", "To'lov sanasi", "Yaratilgan"])
    
    for debt in debts:
        debt_id, employee_name, employee_chat_id, task_id, amount, reason, payment_date, created_at, status = debt
        
        try:
            created_formatted = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
        except:
            created_formatted = created_at
        
        ws3.append([
            debt_id,
            employee_name,
            amount,
            reason,
            payment_date,
            created_formatted
        ])
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"admin_hisobot_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    wb.save(filepath)
    return filepath

def serialize_json_data(data: Dict[str, Any]) -> str:
    """Serialize data to JSON string"""
    return json.dumps(data, ensure_ascii=False)

def parse_json_data(data_str: str) -> Dict[str, Any]:
    """Parse JSON string to data"""
    try:
        return json.loads(data_str) if data_str else {}
    except:
        return {}

def generate_debts_report_excel() -> Optional[str]:
    """Generate Excel report for debts"""
    from database import get_debts
    from datetime import datetime
    import openpyxl
    
    ensure_directories()
    
    # Get all debts
    debts = get_debts()
    if not debts:
        return None
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qarzlar Hisoboti"
    
    # Headers
    headers = [
        "ID", "Xodim", "Miqdor (so'm)", "Sabab", "To'lov sanasi", 
        "Yaratilgan", "Holat"
    ]
    ws.append(headers)
    
    # Data
    total_debt = 0
    paid_debt = 0
    
    for debt in debts:
        debt_id, employee_name, employee_chat_id, task_id, amount, reason, payment_date, created_at, status = debt
        
        if status == 'unpaid':
            total_debt += amount
        else:
            paid_debt += amount
        
        # Format dates
        try:
            created_formatted = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
        except:
            created_formatted = created_at
        
        status_text = "To'lanmagan" if status == 'unpaid' else "To'langan"
        
        ws.append([
            debt_id,
            employee_name,
            amount,
            reason,
            payment_date,
            created_formatted,
            status_text
        ])
    
    # Summary
    ws.append([])
    ws.append(["JAMI:", "", total_debt + paid_debt, "", "", "", ""])
    ws.append(["To'lanmagan:", "", total_debt, "", "", "", ""])
    ws.append(["To'langan:", "", paid_debt, "", "", "", ""])
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"qarzlar_hisoboti_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    wb.save(filepath)
    return filepath

def generate_custom_export(export_type: str) -> Optional[str]:
    """Generate custom data export based on type"""
    ensure_directories()
    
    from database import DATABASE_PATH
    import sqlite3
    
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    wb = openpyxl.Workbook()
    
    try:
        if export_type == "📊 Barcha ma'lumotlar":
            # Tasks sheet
            ws_tasks = wb.active
            ws_tasks.title = "Vazifalar"
            ws_tasks.append(["ID", "Tavsif", "Xodim", "Holat", "To'lov", "Olingan", "Yaratilgan", "Yakunlangan"])
            
            cursor.execute("SELECT * FROM tasks ORDER BY created_at DESC")
            for task in cursor.fetchall():
                ws_tasks.append([
                    task[0], task[1], task[6], task[8], task[5] or 0, 
                    task[14] or 0, task[9], task[11] or ""
                ])
            
            # Debts sheet
            ws_debts = wb.create_sheet("Qarzlar")
            ws_debts.append(["Xodim", "Miqdor", "Sabab", "To'lov sanasi", "Yaratilgan"])
            
            cursor.execute("SELECT * FROM debts ORDER BY created_at DESC")
            for debt in cursor.fetchall():
                ws_debts.append([debt[1], debt[3], debt[4], debt[5], debt[6]])
            
            # Locations sheet
            ws_locations = wb.create_sheet("Lokatsiyalar")
            ws_locations.append(["Xodim", "Latitude", "Longitude", "Tur", "Vaqt"])
            
            cursor.execute("SELECT * FROM employee_locations ORDER BY created_at DESC LIMIT 1000")
            for loc in cursor.fetchall():
                ws_locations.append([loc[1], loc[3], loc[4], loc[5], loc[6]])
        
        elif export_type == "📝 Faqat vazifalar":
            ws = wb.active
            ws.title = "Barcha Vazifalar"
            ws.append(["ID", "Tavsif", "Xodim", "Holat", "To'lov (so'm)", "Olingan (so'm)", 
                      "Joylashuv", "Yaratilgan", "Boshlangan", "Yakunlangan", "Hisobot"])
            
            cursor.execute("SELECT * FROM tasks ORDER BY created_at DESC")
            for task in cursor.fetchall():
                ws.append([
                    task[0], task[1], task[6], task[8], task[5] or 0, task[14] or 0,
                    task[4] or "Belgilanmagan", task[9], task[10] or "", task[11] or "", 
                    task[12][:100] if task[12] else ""
                ])
        
        elif export_type == "💸 Faqat qarzlar":
            ws = wb.active
            ws.title = "Barcha Qarzlar"
            ws.append(["Xodim", "Chat ID", "Vazifa ID", "Miqdor (so'm)", "Sabab", 
                      "To'lov sanasi", "Yaratilgan"])
            
            cursor.execute("SELECT * FROM debts ORDER BY created_at DESC")
            for debt in cursor.fetchall():
                ws.append([debt[1], debt[2], debt[3] or "", debt[4], debt[5], debt[6], debt[7]])
        
        elif export_type == "📍 Lokatsiya tarixi":
            ws = wb.active
            ws.title = "Lokatsiya Tarixi"
            ws.append(["Xodim", "Chat ID", "Latitude", "Longitude", "Tur", "Vaqt", "Live"])
            
            cursor.execute("SELECT * FROM employee_locations ORDER BY created_at DESC")
            for loc in cursor.fetchall():
                ws.append([loc[1], loc[2], loc[3], loc[4], loc[5], loc[6], loc[7]])
        
        conn.close()
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_type = export_type.replace("📊", "").replace("📝", "").replace("💸", "").replace("📍", "").strip()
        filename = f"export_{safe_type.replace(' ', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        
        wb.save(filepath)
        return filepath
        
    except Exception as e:
        print(f"Export error: {e}")
        return None